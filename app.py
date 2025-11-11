"""Streamlit app for FMC to update Excel Key Figure column + Brand/Regional logic."""

from io import BytesIO
from typing import Dict, Iterable, List, Optional, Tuple
import time

import streamlit as st
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException

# ---------------- CONFIG VISUAL ----------------
st.set_page_config(page_title="FMC Automação", page_icon="🌿", layout="centered")

_CUSTOM_CSS = """
<style>
:root {
    --fmc-green: #007a4d;
    --fmc-green-light: #ebf5f1;
}
[data-testid="stAppViewContainer"] {
    background: linear-gradient(180deg, #ffffff 0%, var(--fmc-green-light) 100%);
}
[data-testid="stHeader"] {background: transparent;}
.block-container {padding-top: 3rem; max-width: 880px;}
.stButton button {
    background-color: var(--fmc-green); border: none; color: white; font-weight: 600;
}
.stButton button:hover {background-color: #005f3a;}
.stDownloadButton button {
    background-color: white; border: 2px solid var(--fmc-green);
    color: var(--fmc-green); font-weight: 600;
}
.stDownloadButton button:hover {
    background-color: var(--fmc-green); color: white;
}
</style>
"""
st.markdown(_CUSTOM_CSS, unsafe_allow_html=True)

st.title("🌿 FMC | Portal de Automação de Relatórios")
st.caption("Atualiza colunas 'Key Figure', 'Brand', 'Regional', '25-Feb' e '25-Jan' mantendo o Excel original.")

st.divider()

uploaded_file = st.file_uploader("📂 Envie o arquivo Excel (.xlsx)", type=["xlsx"])

# ---------------- FUNÇÕES AUXILIARES ----------------
def _find_key_figure_headers(sheet: Worksheet) -> List[Tuple[int, int]]:
    """Localiza coordenadas (linha, coluna) onde o header é 'Key Figure'."""
    headers = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            if isinstance(cell.value, str) and cell.value.strip().lower() == "key figure":
                headers.append((cell.row, cell.column))
    return headers


def _replace_input_values(sheet: Worksheet, headers: Iterable[Tuple[int, int]]) -> int:
    """Substitui 'Input Nov' por 'Input Dez' abaixo de cada header 'Key Figure'."""
    replacements = 0
    for header_row, column in headers:
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=column)
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and "Input Nov" in cell.value:
                cell.value = cell.value.replace("Input Nov", "Input Dez")
                replacements += 1
    return replacements


def _process_workbook(xlsx_bytes: bytes) -> Tuple[BytesIO, Dict[str, float]]:
    """Executa todas as automações e retorna o arquivo atualizado + tempos."""
    t_start = time.time()
    timings = {}
    workbook = load_workbook(BytesIO(xlsx_bytes), data_only=False)
    total_changes = 0
    changed_rows = set()

    # 1️⃣ Input Nov → Input Dez
    t1 = time.time()
    for sheet in workbook.worksheets:
        headers = _find_key_figure_headers(sheet)
        if headers:
            total_changes += _replace_input_values(sheet, headers)
    timings["Input Nov → Dez"] = time.time() - t1

    # 2️⃣ Alterações em Brand e Regional (apenas na aba 'Base' se existir)
    if "Base" in workbook.sheetnames:
        sheet = workbook["Base"]

        # Encontrar cabeçalhos
        header_cells = list(next(sheet.iter_rows(min_row=1, max_row=1)))
        header_map = {cell.value: cell.column for cell in header_cells if cell.value}

        col_brand = header_map.get("Brand")
        col_regional = header_map.get("Regional")
        col_25feb = header_map.get("25-Feb")
        col_25jan = header_map.get("25-Jan")

        # Brand e Regional
        t2 = time.time()
        if col_brand:
            for row in range(2, sheet.max_row + 1):
                val = sheet.cell(row=row, column=col_brand).value
                if val == "ALTACOR":
                    sheet.cell(row=row, column=col_brand).value = "B1"
                    changed_rows.add(row)
                    total_changes += 1
                elif val == "AMETISTA":
                    sheet.cell(row=row, column=col_brand).value = "B2"
                    changed_rows.add(row)
                    total_changes += 1

        if col_regional:
            for row in range(2, sheet.max_row + 1):
                val = sheet.cell(row=row, column=col_regional).value
                if val == "Cana Cerrado":
                    sheet.cell(row=row, column=col_regional).value = "B3"
                    changed_rows.add(row)
                    total_changes += 1
        timings["Brand/Regional"] = time.time() - t2

        # 3️⃣ Atualizar 25-Feb e 25-Jan apenas nas linhas alteradas
        t3 = time.time()
        for row in changed_rows:
            if col_25feb:
                sheet.cell(row=row, column=col_25feb).value = 10
                total_changes += 1
            if col_25jan:
                sheet.cell(row=row, column=col_25jan).value = 10
                total_changes += 1
        timings["25-Feb/25-Jan"] = time.time() - t3

    # 4️⃣ Salvar workbook
    t4 = time.time()
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    timings["Salvar Excel"] = time.time() - t4

    timings["Tempo total"] = time.time() - t_start
    timings["Alterações totais"] = total_changes

    return output, timings

# ---------------- EXECUÇÃO PRINCIPAL ----------------
if uploaded_file:
    try:
        result_stream, timing_report = _process_workbook(uploaded_file.getvalue())

        st.success(f"✅ Automação concluída com {timing_report['Alterações totais']} alterações!")
        st.write("**⏱️ Tempos de execução por etapa:**")
        for etapa, tempo in timing_report.items():
            if etapa not in ("Alterações totais",):
                st.write(f"• {etapa}: {tempo:.2f} s")

        st.download_button(
            "📥 Baixar Excel Atualizado",
            data=result_stream.getvalue(),
            file_name="fmc_atualizado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except InvalidFileException:
        st.error("❌ Arquivo inválido. Verifique se é um .xlsx compatível.")
    except Exception as e:
        st.error(f"❌ Erro ao processar o arquivo: {e}")
